"""
Exportadores de Reportes - Arte Ideas Analytics
Funcionalidad para exportar reportes a Excel y PDF
"""
from io import BytesIO
from datetime import datetime
from decimal import Decimal
import json


class ExcelExporter:
    """Exportador de reportes a Excel"""
    
    def __init__(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            self.Workbook = Workbook
            self.Font = Font
            self.PatternFill = PatternFill
            self.Alignment = Alignment
            self.Border = Border
            self.Side = Side
            self.get_column_letter = get_column_letter
            self.openpyxl_available = True
        except ImportError:
            self.openpyxl_available = False
            # Fallback a xlsxwriter si openpyxl no está disponible
            try:
                import xlsxwriter
                self.xlsxwriter = xlsxwriter
                self.xlsxwriter_available = True
            except ImportError:
                self.xlsxwriter_available = False
    
    def export_report(self, titulo, metricas, detalle, fecha_inicio, fecha_fin, categoria):
        """
        Exportar reporte a Excel
        Retorna bytes del archivo Excel
        """
        if self.openpyxl_available:
            return self._export_with_openpyxl(titulo, metricas, detalle, fecha_inicio, fecha_fin, categoria)
        elif self.xlsxwriter_available:
            return self._export_with_xlsxwriter(titulo, metricas, detalle, fecha_inicio, fecha_fin, categoria)
        else:
            raise ImportError("Se requiere openpyxl o xlsxwriter para exportar a Excel. Instale con: pip install openpyxl")
    
    def _export_with_openpyxl(self, titulo, metricas, detalle, fecha_inicio, fecha_fin, categoria):
        """Exportar usando openpyxl"""
        wb = self.Workbook()
        ws = wb.active
        ws.title = "Reporte"
        
        # Estilos
        header_fill = self.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = self.Font(bold=True, color="FFFFFF", size=12)
        title_font = self.Font(bold=True, size=14)
        metric_font = self.Font(bold=True, size=11)
        
        border = self.Border(
            left=self.Side(style='thin'),
            right=self.Side(style='thin'),
            top=self.Side(style='thin'),
            bottom=self.Side(style='thin')
        )
        
        # Título
        row = 1
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = titulo
        cell.font = title_font
        cell.alignment = self.Alignment(horizontal='center', vertical='center')
        
        # Información del período
        row += 2
        ws[f'A{row}'] = "Período:"
        ws[f'B{row}'] = f"{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        ws[f'A{row}'].font = self.Font(bold=True)
        
        ws[f'C{row}'] = "Fecha de Generación:"
        ws[f'D{row}'] = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        ws[f'C{row}'].font = self.Font(bold=True)
        
        # Métricas
        row += 2
        ws[f'A{row}'] = "MÉTRICAS DE RESUMEN"
        ws[f'A{row}'].font = metric_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].font = header_font
        
        row += 1
        metric_col = 0
        for key, value in metricas.items():
            if metric_col >= 4:  # Máximo 4 columnas
                metric_col = 0
                row += 1
            
            col_letter = self.get_column_letter(metric_col + 1)
            ws[f'{col_letter}{row}'] = self._format_metric_label(key)
            ws[f'{col_letter}{row}'].font = self.Font(bold=True)
            
            col_value_letter = self.get_column_letter(metric_col + 2)
            ws[f'{col_value_letter}{row}'] = self._format_metric_value(value)
            
            metric_col += 2
        
        # Tabla de detalle
        if detalle:
            row += 2
            ws[f'A{row}'] = "DETALLE"
            ws[f'A{row}'].font = metric_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].font = header_font
            
            row += 1
            
            # Encabezados
            if detalle:
                headers = list(detalle[0].keys())
                for col_idx, header in enumerate(headers):
                    col_letter = self.get_column_letter(col_idx + 1)
                    cell = ws[f'{col_letter}{row}']
                    cell.value = self._format_header(header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = self.Alignment(horizontal='center', vertical='center')
                    cell.border = border
                
                # Datos
                row += 1
                for item in detalle:
                    for col_idx, header in enumerate(headers):
                        col_letter = self.get_column_letter(col_idx + 1)
                        cell = ws[f'{col_letter}{row}']
                        value = item.get(header, '')
                        cell.value = self._format_cell_value(value)
                        cell.border = border
                    row += 1
                
                # Ajustar ancho de columnas
                for col_idx, header in enumerate(headers):
                    col_letter = self.get_column_letter(col_idx + 1)
                    ws.column_dimensions[col_letter].width = 20
        
        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _export_with_xlsxwriter(self, titulo, metricas, detalle, fecha_inicio, fecha_fin, categoria):
        """Exportar usando xlsxwriter (fallback)"""
        output = BytesIO()
        workbook = self.xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Reporte')
        
        # Formatos
        header_format = workbook.add_format({
            'bold': True,
            'font_color': 'white',
            'bg_color': '#366092',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        title_format = workbook.add_format({
            'bold': True,
            'size': 14,
            'align': 'center'
        })
        
        metric_label_format = workbook.add_format({'bold': True})
        
        # Título
        worksheet.merge_range('A1:D1', titulo, title_format)
        
        # Período
        row = 2
        worksheet.write(row, 0, 'Período:', metric_label_format)
        worksheet.write(row, 1, f"{fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}")
        worksheet.write(row, 2, 'Fecha de Generación:', metric_label_format)
        worksheet.write(row, 3, datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
        
        # Métricas
        row = 4
        worksheet.write(row, 0, 'MÉTRICAS DE RESUMEN', header_format)
        row += 1
        
        metric_col = 0
        for key, value in metricas.items():
            if metric_col >= 4:
                metric_col = 0
                row += 1
            
            worksheet.write(row, metric_col, self._format_metric_label(key), metric_label_format)
            worksheet.write(row, metric_col + 1, self._format_metric_value(value))
            metric_col += 2
        
        # Tabla de detalle
        if detalle:
            row += 2
            worksheet.write(row, 0, 'DETALLE', header_format)
            row += 1
            
            headers = list(detalle[0].keys())
            for col_idx, header in enumerate(headers):
                worksheet.write(row, col_idx, self._format_header(header), header_format)
            
            row += 1
            for item in detalle:
                for col_idx, header in enumerate(headers):
                    value = item.get(header, '')
                    worksheet.write(row, col_idx, self._format_cell_value(value))
                row += 1
            
            # Ajustar ancho de columnas
            for col_idx in range(len(headers)):
                worksheet.set_column(col_idx, col_idx, 20)
        
        workbook.close()
        output.seek(0)
        return output.getvalue()
    
    def _format_metric_label(self, key):
        """Formatear etiqueta de métrica"""
        # Convertir snake_case a Title Case
        return key.replace('_', ' ').title()
    
    def _format_metric_value(self, value):
        """Formatear valor de métrica"""
        if isinstance(value, (int, float, Decimal)):
            if isinstance(value, float) and value.is_integer():
                return int(value)
            return value
        return str(value)
    
    def _format_header(self, header):
        """Formatear encabezado de columna"""
        return header.replace('_', ' ').title()
    
    def _format_cell_value(self, value):
        """Formatear valor de celda"""
        if value is None:
            return ''
        if isinstance(value, bool):
            return 'Sí' if value else 'No'
        if isinstance(value, (int, float, Decimal)):
            return value
        return str(value)


class PDFExporter:
    """Exportador de reportes a PDF"""
    
    def __init__(self):
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            
            self.colors = colors
            self.letter = letter
            self.A4 = A4
            self.SimpleDocTemplate = SimpleDocTemplate
            self.Table = Table
            self.TableStyle = TableStyle
            self.Paragraph = Paragraph
            self.Spacer = Spacer
            self.getSampleStyleSheet = getSampleStyleSheet
            self.ParagraphStyle = ParagraphStyle
            self.inch = inch
            self.reportlab_available = True
        except ImportError:
            self.reportlab_available = False
    
    def export_report(self, titulo, metricas, detalle, fecha_inicio, fecha_fin, categoria):
        """
        Exportar reporte a PDF
        Retorna bytes del archivo PDF
        """
        if not self.reportlab_available:
            raise ImportError("Se requiere reportlab para exportar a PDF. Instale con: pip install reportlab")
        
        return self._export_with_reportlab(titulo, metricas, detalle, fecha_inicio, fecha_fin, categoria)
    
    def _export_with_reportlab(self, titulo, metricas, detalle, fecha_inicio, fecha_fin, categoria):
        """Exportar usando reportlab"""
        buffer = BytesIO()
        
        # Crear documento
        doc = self.SimpleDocTemplate(buffer, pagesize=self.A4)
        elements = []
        
        # Estilos
        styles = self.getSampleStyleSheet()
        title_style = self.ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=self.colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1  # Centrado
        )
        
        # Título
        title = self.Paragraph(titulo, title_style)
        elements.append(title)
        elements.append(self.Spacer(1, 0.2 * self.inch))
        
        # Información del período
        periodo_text = f"<b>Período:</b> {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        fecha_gen_text = f"<b>Fecha de Generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        
        periodo_para = self.Paragraph(periodo_text, styles['Normal'])
        fecha_para = self.Paragraph(fecha_gen_text, styles['Normal'])
        
        elements.append(periodo_para)
        elements.append(fecha_para)
        elements.append(self.Spacer(1, 0.3 * self.inch))
        
        # Métricas
        elements.append(self.Paragraph("<b>MÉTRICAS DE RESUMEN</b>", styles['Heading2']))
        elements.append(self.Spacer(1, 0.1 * self.inch))
        
        # Crear tabla de métricas
        metric_data = []
        metric_items = list(metricas.items())
        
        # Organizar métricas en filas de 2 columnas
        for i in range(0, len(metric_items), 2):
            row = []
            row.append(self._format_metric_label(metric_items[i][0]))
            row.append(str(self._format_metric_value(metric_items[i][1])))
            
            if i + 1 < len(metric_items):
                row.append(self._format_metric_label(metric_items[i + 1][0]))
                row.append(str(self._format_metric_value(metric_items[i + 1][1])))
            else:
                row.append('')
                row.append('')
            
            metric_data.append(row)
        
        metric_table = self.Table(metric_data, colWidths=[2 * self.inch, 2 * self.inch, 2 * self.inch, 2 * self.inch])
        metric_table.setStyle(self.TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), self.colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), self.colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, self.colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(metric_table)
        elements.append(self.Spacer(1, 0.3 * self.inch))
        
        # Tabla de detalle
        if detalle:
            elements.append(self.Paragraph("<b>DETALLE</b>", styles['Heading2']))
            elements.append(self.Spacer(1, 0.1 * self.inch))
            
            # Preparar datos de la tabla
            headers = list(detalle[0].keys())
            table_data = [headers]
            
            for item in detalle:
                row = [str(self._format_cell_value(item.get(header, ''))) for header in headers]
                table_data.append(row)
            
            # Crear tabla
            # Ajustar ancho de columnas según cantidad
            num_cols = len(headers)
            col_width = 7 * self.inch / num_cols if num_cols > 0 else 2 * self.inch
            col_widths = [col_width] * num_cols
            
            detail_table = self.Table(table_data, colWidths=col_widths)
            detail_table.setStyle(self.TableStyle([
                # Encabezado
                ('BACKGROUND', (0, 0), (-1, 0), self.colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), self.colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                # Datos
                ('BACKGROUND', (0, 1), (-1, -1), self.colors.white),
                ('TEXTCOLOR', (0, 1), (-1, -1), self.colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, self.colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                # Alternar colores de filas
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [self.colors.white, self.colors.lightgrey]),
            ]))
            
            elements.append(detail_table)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
    
    def _format_metric_label(self, key):
        """Formatear etiqueta de métrica"""
        return key.replace('_', ' ').title()
    
    def _format_metric_value(self, value):
        """Formatear valor de métrica"""
        if isinstance(value, (int, float, Decimal)):
            if isinstance(value, float) and value.is_integer():
                return int(value)
            return value
        return str(value)
    
    def _format_cell_value(self, value):
        """Formatear valor de celda"""
        if value is None:
            return ''
        if isinstance(value, bool):
            return 'Sí' if value else 'No'
        if isinstance(value, (int, float, Decimal)):
            return value
        return str(value)

